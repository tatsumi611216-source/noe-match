#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""同棲〜結婚〜結婚式〜妊活〜出産〜子1歳 のロードマップ／WBS／リスク管理表を
1つのスプレッドシートとして生成する。

データは scripts/wbs_data.py に分離。

usage: python3 scripts/build_life_wbs_xlsx.py
out:   docs/同棲から子育てまでのWBS.xlsx
"""

import datetime
import os
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wbs_data import PHASES, RISKS, TASKS  # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "docs", "同棲から子育てまでのWBS.xlsx")

FONT = "Meiryo"
BASE = "'はじめに'!$C$5"
GANTT_MONTHS = 49

INK = "1F2A24"
ACCENT = "2F6B4F"
CRIT = "A93C27"

F_HEAD = PatternFill("solid", fgColor="2F6B4F")
F_HEAD_CRIT = PatternFill("solid", fgColor="A93C27")
F_INPUT = PatternFill("solid", fgColor="FFF6CC")
F_LV2 = PatternFill("solid", fgColor="E6EDE7")
F_BAND = PatternFill("solid", fgColor="F5F7F4")
F_BAR_PRE = PatternFill("solid", fgColor="E5D3AE")
F_BAR_POST = PatternFill("solid", fgColor="BFD9C6")
F_TOTAL = PatternFill("solid", fgColor="E6ECE6")

THIN = Side(style="thin", color="C9D2C9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS = '"未着手,進行中,完了,対象外"'


def head_font():
    return Font(name=FONT, size=10, bold=True, color="FFFFFF")


def body_font(bold=False, color=INK, size=10):
    return Font(name=FONT, size=size, bold=bold, color=color)


def style_header(ws, row, ncols, fill):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = head_font()
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 36


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def add_status_cf(ws, ref):
    ws.conditional_formatting.add(ref, CellIsRule(
        operator="equal", formula=['"完了"'],
        fill=PatternFill("solid", fgColor="D6EBDC"), font=Font(name=FONT, color=ACCENT)))
    ws.conditional_formatting.add(ref, CellIsRule(
        operator="equal", formula=['"進行中"'],
        fill=PatternFill("solid", fgColor="FCEFD0"), font=Font(name=FONT, color="8A6428")))
    ws.conditional_formatting.add(ref, CellIsRule(
        operator="equal", formula=['"対象外"'],
        fill=PatternFill("solid", fgColor="EDEFEC"), font=Font(name=FONT, color="9AA69C")))


# ===========================================================================
# ① ロードマップ
# ===========================================================================

def sheet_roadmap(wb):
    ws = wb.create_sheet("①ロードマップ")
    ws.cell(row=1, column=1,
            value="① ロードマップ｜同棲 → 結婚 → 結婚式 → 妊活 → 出産 → 子1歳").font = \
        Font(name=FONT, size=14, bold=True, color=ACCENT)
    ws.cell(row=1, column=5,
            value="黄色のセルを書き換えると、日付とガントが自動で動きます").font = \
        Font(name=FONT, size=9, color="6B756E")
    ws.row_dimensions[1].height = 24

    heads = ["ID", "フェーズ", "このフェーズのゴール（完了条件）", "期間\n(月)",
             "開始\n(経過月)", "終了\n(経過月)", "予定開始", "予定完了",
             "主要マイルストーン", "このフェーズで決めること（意思決定）",
             "費用下限\n(万円)", "費用上限\n(万円)", "想定される最大リスク", "状態"]
    ncol = len(heads)

    ws.cell(row=2, column=ncol, value="経過月 →").font = body_font(size=8, color="6B756E")
    ws.cell(row=2, column=ncol).alignment = Alignment(horizontal="right")
    for k in range(GANTT_MONTHS):
        c = ws.cell(row=2, column=ncol + 1 + k, value=k)
        c.font = Font(name=FONT, size=7, color="9AA69C")
        c.alignment = Alignment(horizontal="center")

    for i, h in enumerate(heads, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, ncol, F_HEAD)

    for k in range(GANTT_MONTHS):
        col = ncol + 1 + k
        c = ws.cell(row=3, column=col,
                    value="=EDATE(%s,%s2)" % (BASE, get_column_letter(col)))
        c.number_format = 'yy"/"m'
        c.font = Font(name=FONT, size=7, bold=True, color="FFFFFF")
        c.fill = F_HEAD
        c.alignment = Alignment(textRotation=90, horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 3.0

    r = 4
    for pid, name, goal, dur, start, ms, decide, lo, hi, risk in PHASES:
        ws.cell(row=r, column=1, value=pid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=goal)
        ws.cell(row=r, column=4, value=dur).fill = F_INPUT
        ws.cell(row=r, column=5, value=start).fill = F_INPUT
        ws.cell(row=r, column=6, value="=E%d+D%d-1" % (r, r))
        g = ws.cell(row=r, column=7, value="=EDATE(%s,E%d)" % (BASE, r))
        h = ws.cell(row=r, column=8, value="=EDATE(%s,F%d)" % (BASE, r))
        g.number_format = h.number_format = 'yyyy"年"m"月"'
        ws.cell(row=r, column=9, value=ms)
        ws.cell(row=r, column=10, value=decide)
        ws.cell(row=r, column=11, value=lo).fill = F_INPUT
        ws.cell(row=r, column=12, value=hi).fill = F_INPUT
        ws.cell(row=r, column=13, value=risk)
        ws.cell(row=r, column=14, value="未着手").fill = F_INPUT
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font(bold=(c == 2))
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 9, 10, 13)))
            cell.border = BOX
        for c in (4, 5, 6, 11, 12):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="top")
            ws.cell(row=r, column=c).number_format = "#,##0"
        ws.row_dimensions[r].height = 58
        r += 1

    last = r - 1
    ws.cell(row=r, column=2, value="合計").font = body_font(bold=True)
    ws.cell(row=r, column=11, value="=SUM(K4:K%d)" % last).font = body_font(bold=True)
    ws.cell(row=r, column=12, value="=SUM(L4:L%d)" % last).font = body_font(bold=True)
    ws.cell(row=r, column=13,
            value="※ ご祝儀・親からの援助・各種給付金は差し引く前の総額").font = \
        body_font(size=9, color="6B756E")
    for c in range(1, ncol + 1):
        ws.cell(row=r, column=c).fill = F_TOTAL
        ws.cell(row=r, column=c).border = BOX
    for c in (11, 12):
        ws.cell(row=r, column=c).number_format = "#,##0"
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")

    g0 = get_column_letter(ncol + 1)
    g1 = get_column_letter(ncol + GANTT_MONTHS)
    ws.conditional_formatting.add(
        "%s4:%s11" % (g0, g1),
        FormulaRule(formula=["AND(%s$2>=$E4,%s$2<=$F4)" % (g0, g0)],
                    fill=F_BAR_PRE, stopIfTrue=True))
    ws.conditional_formatting.add(
        "%s4:%s%d" % (g0, g1, last),
        FormulaRule(formula=["AND(%s$2>=$E4,%s$2<=$F4)" % (g0, g0)],
                    fill=F_BAR_POST, stopIfTrue=True))

    set_widths(ws, {"A": 5, "B": 20, "C": 34, "D": 6, "E": 7, "F": 7, "G": 12,
                    "H": 12, "I": 28, "J": 44, "K": 9, "L": 9, "M": 34, "N": 9})

    dv = DataValidation(type="list", formula1=STATUS, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("N4:N%d" % last)
    add_status_cf(ws, "N4:N%d" % last)

    ws.freeze_panes = "%s4" % g0
    ws.auto_filter.ref = "A3:%s%d" % (get_column_letter(ncol), last)
    ws.sheet_view.zoomScale = 90


# ===========================================================================
# ② WBS
# ===========================================================================

def sheet_wbs(wb):
    ws = wb.create_sheet("②WBS")
    ws.cell(row=1, column=1,
            value="② WBS｜想定される論点まで分解したタスク一覧").font = \
        Font(name=FONT, size=14, bold=True, color=ACCENT)
    ws.cell(row=1, column=5,
            value="濃い行＝タスク／薄い行＝その中の実行ステップ。左の＋−で折りたためます").font = \
        Font(name=FONT, size=9, color="6B756E")
    ws.row_dimensions[1].height = 24

    heads = ["WBS-ID", "フェーズ", "中分類", "Lv", "タスク／実行ステップ",
             "完了の定義（これができたら完了）", "論点・決めること",
             "選択肢・候補", "判断軸・注意点", "経過\n月", "目安時期",
             "区分", "担当", "所要\n目安", "費用目安",
             "先行して終わっていること", "状態", "メモ"]
    for i, h in enumerate(heads, start=1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(heads), F_HEAD)

    ws.sheet_properties.outlinePr.summaryBelow = False

    major = {}
    minor = 0
    cur_id = ""
    r = 3
    for (ph, cat, lv, task, dod, m, kind, who, eff, cost, dep,
         ronten, opts, axis) in TASKS:
        if lv == 2:
            major[ph] = major.get(ph, 0) + 1
            minor = 0
            cur_id = "%s-%02d" % (ph, major[ph])
            wid = cur_id
        else:
            minor += 1
            wid = "%s-%02d" % (cur_id, minor)

        ws.cell(row=r, column=1, value=wid)
        ws.cell(row=r, column=2, value=ph)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=lv)
        ws.cell(row=r, column=5, value=task)
        ws.cell(row=r, column=6, value=dod)
        ws.cell(row=r, column=7, value=ronten)
        ws.cell(row=r, column=8, value=opts)
        ws.cell(row=r, column=9, value=axis)
        ws.cell(row=r, column=10, value=m)
        d = ws.cell(row=r, column=11, value="=EDATE(%s,J%d)" % (BASE, r))
        d.number_format = 'yyyy"/"m'
        ws.cell(row=r, column=12, value=kind)
        ws.cell(row=r, column=13, value=who)
        ws.cell(row=r, column=14, value=eff)
        ws.cell(row=r, column=15, value=cost)
        ws.cell(row=r, column=16, value=dep)
        ws.cell(row=r, column=17, value="未着手").fill = F_INPUT
        ws.cell(row=r, column=18, value="").fill = F_INPUT

        for c in range(1, len(heads) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font(bold=(lv == 2 and c == 5))
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(c in (5, 6, 7, 8, 9, 15, 16)),
                indent=(2 if (lv == 3 and c == 5) else 0))
            cell.border = BOX
        for c in (1, 2, 4, 10, 11, 12, 13, 14, 17):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="top")

        if lv == 2:
            for c in range(1, len(heads) + 1):
                if c not in (17, 18):
                    ws.cell(row=r, column=c).fill = F_LV2
        else:
            ws.row_dimensions[r].outline_level = 1
            if r % 2 == 1:
                for c in range(1, len(heads) + 1):
                    if c not in (17, 18):
                        ws.cell(row=r, column=c).fill = F_BAND
        r += 1

    last = r - 1
    set_widths(ws, {"A": 12, "B": 7, "C": 9, "D": 4, "E": 44, "F": 34, "G": 26,
                    "H": 50, "I": 54, "J": 6, "K": 10, "L": 7, "M": 13,
                    "N": 8, "O": 14, "P": 22, "Q": 9, "R": 20})

    dv = DataValidation(type="list", formula1=STATUS, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("Q3:Q%d" % last)
    add_status_cf(ws, "Q3:Q%d" % last)

    ws.conditional_formatting.add("L3:L%d" % last, CellIsRule(
        operator="equal", formula=['"法定"'],
        fill=PatternFill("solid", fgColor="F6E3DD"),
        font=Font(name=FONT, bold=True, color=CRIT)))

    ws.freeze_panes = "F3"
    ws.auto_filter.ref = "A2:R%d" % last
    ws.sheet_view.zoomScale = 85
    return last


# ===========================================================================
# ③ リスク管理表
# ===========================================================================

def sheet_risk(wb):
    ws = wb.create_sheet("③リスク管理表")
    ws.cell(row=1, column=1,
            value="③ リスク管理表｜各フェーズで配偶者との関係が悪化する想定リスクと対策").font = \
        Font(name=FONT, size=14, bold=True, color=CRIT)
    ws.cell(row=1, column=6,
            value="確率×影響でスコアと優先度が自動計算されます").font = \
        Font(name=FONT, size=9, color="6B756E")
    ws.row_dimensions[1].height = 24

    heads = ["リスクID", "フェーズ", "分類", "リスク事象", "典型的な場面",
             "早期サイン（黄信号）", "相手側から見えている景色",
             "確率\n(1-5)", "影響\n(1-5)", "スコア", "優先度",
             "予防策（起きる前にやること）", "発生時の初動（24時間以内）",
             "再発防止（構造の変更）", "オーナー", "見直し", "状態"]
    for i, h in enumerate(heads, start=1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(heads), F_HEAD_CRIT)

    r = 3
    for i, (ph, cat, ev, scene, sign, view, p, imp, prev, first, fix, owner) \
            in enumerate(RISKS, start=1):
        ws.cell(row=r, column=1, value="R%03d" % i)
        ws.cell(row=r, column=2, value=ph)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=ev)
        ws.cell(row=r, column=5, value=scene)
        ws.cell(row=r, column=6, value=sign)
        ws.cell(row=r, column=7, value=view)
        ws.cell(row=r, column=8, value=p).fill = F_INPUT
        ws.cell(row=r, column=9, value=imp).fill = F_INPUT
        ws.cell(row=r, column=10, value="=H%d*I%d" % (r, r))
        ws.cell(row=r, column=11,
                value='=IF(J%d>=16,"最優先",IF(J%d>=12,"高",IF(J%d>=6,"中","低")))'
                      % (r, r, r))
        ws.cell(row=r, column=12, value=prev)
        ws.cell(row=r, column=13, value=first)
        ws.cell(row=r, column=14, value=fix)
        ws.cell(row=r, column=15, value=owner)
        ws.cell(row=r, column=16, value="月1回")
        ws.cell(row=r, column=17, value="未着手").fill = F_INPUT

        for c in range(1, len(heads) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = body_font(bold=(c == 4))
            cell.alignment = Alignment(
                vertical="top", wrap_text=(c in (4, 5, 6, 7, 12, 13, 14)))
            cell.border = BOX
        for c in (1, 2, 3, 8, 9, 10, 11, 15, 16, 17):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="top")
        if r % 2 == 1:
            for c in range(1, len(heads) + 1):
                if c not in (8, 9, 17):
                    ws.cell(row=r, column=c).fill = F_BAND
        r += 1

    last = r - 1
    set_widths(ws, {"A": 8, "B": 7, "C": 12, "D": 32, "E": 24, "F": 26, "G": 34,
                    "H": 6, "I": 6, "J": 7, "K": 8, "L": 52, "M": 40, "N": 32,
                    "O": 12, "P": 8, "Q": 9})

    dv = DataValidation(type="list", formula1=STATUS, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("Q3:Q%d" % last)
    add_status_cf(ws, "Q3:Q%d" % last)

    dv2 = DataValidation(type="whole", operator="between", formula1="1",
                         formula2="5", allow_blank=True, showErrorMessage=True)
    dv2.error = "1〜5の整数を入れてください"
    ws.add_data_validation(dv2)
    dv2.add("H3:I%d" % last)

    ws.conditional_formatting.add("K3:K%d" % last, CellIsRule(
        operator="equal", formula=['"最優先"'],
        fill=PatternFill("solid", fgColor="A93C27"),
        font=Font(name=FONT, bold=True, color="FFFFFF")))
    ws.conditional_formatting.add("K3:K%d" % last, CellIsRule(
        operator="equal", formula=['"高"'],
        fill=PatternFill("solid", fgColor="F0C9BE"), font=Font(name=FONT, color=CRIT)))
    ws.conditional_formatting.add("K3:K%d" % last, CellIsRule(
        operator="equal", formula=['"中"'],
        fill=PatternFill("solid", fgColor="FCEFD0"), font=Font(name=FONT, color="8A6428")))
    ws.conditional_formatting.add("J3:J%d" % last, CellIsRule(
        operator="greaterThanOrEqual", formula=["16"],
        font=Font(name=FONT, bold=True, color=CRIT)))

    ws.freeze_panes = "D3"
    ws.auto_filter.ref = "A2:Q%d" % last
    ws.sheet_view.zoomScale = 85
    return last


# ===========================================================================
# はじめに
# ===========================================================================

def sheet_intro(wb, wbs_last, risk_last):
    ws = wb.create_sheet("はじめに", 0)
    ws.cell(row=1, column=1,
            value="同棲 → 結婚 → 結婚式 → 妊活 → 出産 → 子1歳 の計画表").font = \
        Font(name=FONT, size=16, bold=True, color=ACCENT)
    ws.cell(row=2, column=1,
            value="作成日：2026-08-11 ／ 制度情報は2026年8月時点").font = \
        Font(name=FONT, size=9, color="6B756E")

    ws.cell(row=4, column=1, value="■ 最初にここだけ入力してください").font = \
        body_font(bold=True, size=11)
    ws.cell(row=5, column=1,
            value="基準日（＝計画の起点。同棲を始める月の1日を入れる）").font = body_font()
    base = ws.cell(row=5, column=3, value=datetime.date(2026, 10, 1))
    base.number_format = "yyyy/mm/dd"
    base.fill = F_INPUT
    base.font = Font(name=FONT, size=11, bold=True, color="0000FF")
    base.border = BOX
    ws.cell(row=6, column=1,
            value="→ この日付を変えると、①②の「予定開始／予定完了／目安時期」とガントが全部動きます").font = \
        body_font(size=9, color="6B756E")

    ws.cell(row=8, column=1, value="■ シートの構成").font = body_font(bold=True, size=11)
    rows = [
        ("①ロードマップ", "13フェーズの全体像。期間・費用・意思決定・リスクを1行にまとめ、右側に49ヶ月分のガントチャート"),
        ("②WBS", "①を分解した実行タスク一覧。濃い行＝タスク、薄い行＝その中の実行ステップ。"
                 "「論点・決めること」「選択肢・候補」「判断軸・注意点」の3列が実際に迷う場所"),
        ("③リスク管理表", "各フェーズで配偶者との関係が悪化する想定リスクと、"
                      "早期サイン・相手側から見えている景色・予防策・初動・再発防止"),
    ]
    r = 9
    for name, desc in rows:
        ws.cell(row=r, column=1, value=name).font = body_font(bold=True, color=ACCENT)
        c = ws.cell(row=r, column=3, value=desc)
        c.font = body_font()
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 34
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ 使い方のコツ").font = body_font(bold=True, size=11)
    r += 1
    tips = [
        "②WBS は行がグループ化してあります。シート左の「−」を押すと実行ステップが折りたたまれ、タスクだけの一覧になります。",
        "②WBS の2行目でフィルタをかけられます。「フェーズ」で今の時期だけ、「区分＝法定」で期限のあるものだけ、「担当」で自分の分だけ、という見方ができます。",
        "③リスク管理表は、確率と影響を自分たちの実感に合わせて1〜5で書き換えてください。スコアと優先度が自動で変わります。",
        "毎週15分のミーティングで、②の「進行中」の行と③の「早期サイン」列だけを2人で見る、という使い方を想定しています。",
    ]
    for tip in tips:
        c = ws.cell(row=r, column=1, value=tip)
        c.font = body_font(size=9, color="4B564E")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ 色と表記の意味").font = body_font(bold=True, size=11)
    r += 1
    legend = [
        ("黄色のセル", "自分たちで書き換える入力欄（基準日・期間・費用・確率・影響・状態・メモ）"),
        ("白いセル", "そのまま使う内容、または自動計算（数式が入っているので上書きしない）"),
        ("赤い文字・赤い背景", "法定期限があるもの／リスクの優先度が高いもの"),
        ("担当欄の「妊娠する側」",
         "妊娠・出産する本人。「パートナー」はもう一方。あくまで一般的な例なので、実際の家庭に合わせて書き換えてください"),
        ("Lv列", "2＝タスク（意思決定を含む親）、3＝その中の実行ステップ"),
    ]
    for name, desc in legend:
        c = ws.cell(row=r, column=1, value=name)
        c.font = body_font(bold=True)
        if name == "黄色のセル":
            c.fill = F_INPUT
        c.border = BOX
        d = ws.cell(row=r, column=3, value=desc)
        d.font = body_font()
        d.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ 記入例（②WBS の使い方）").font = \
        body_font(bold=True, size=11)
    r += 1
    for i, h in enumerate(["WBS-ID", "タスク", "状態", "メモ"], start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = F_HEAD
        cell.font = head_font()
        cell.border = BOX
    r += 1
    ex = ["P1-02-03", "内定最低指数を調べる", "進行中",
          "9/20に◯◯区と△△市を比較。10/5までに結論"]
    for i, v in enumerate(ex, start=1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = body_font()
        cell.border = BOX
        if i >= 3:
            cell.fill = F_INPUT
    r += 2

    ws.cell(row=r, column=1, value="■ 進捗サマリー（自動集計）").font = \
        body_font(bold=True, size=11)
    r += 1
    for i, h in enumerate(["", "全件", "未着手", "進行中", "完了", "対象外", "進捗率"],
                          start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.fill = F_HEAD
        cell.font = head_font()
        cell.alignment = Alignment(horizontal="center")
        cell.border = BOX
    r += 1

    for label, sheet, col, lastrow in (
            ("②WBS", "'②WBS'", "Q", wbs_last),
            ("③リスク管理表", "'③リスク管理表'", "Q", risk_last),
    ):
        rng = "%s!$%s$3:$%s$%d" % (sheet, col, col, lastrow)
        ws.cell(row=r, column=1, value=label).font = body_font(bold=True)
        ws.cell(row=r, column=2, value="=COUNTA(%s)" % rng)
        ws.cell(row=r, column=3, value='=COUNTIF(%s,"未着手")' % rng)
        ws.cell(row=r, column=4, value='=COUNTIF(%s,"進行中")' % rng)
        ws.cell(row=r, column=5, value='=COUNTIF(%s,"完了")' % rng)
        ws.cell(row=r, column=6, value='=COUNTIF(%s,"対象外")' % rng)
        pc = ws.cell(row=r, column=7,
                     value="=IF(B%d-F%d=0,0,E%d/(B%d-F%d))" % (r, r, r, r, r))
        pc.number_format = "0.0%"
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            if c > 1:
                cell.font = body_font()
                cell.alignment = Alignment(horizontal="center")
            cell.border = BOX
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ 前提と注意").font = body_font(bold=True, size=11)
    r += 1
    notes = [
        "・フェーズの期間はあくまで標準モデル。特に妊活（P8）は個人差が大きく、数ヶ月〜数年の幅がある。①の黄色いセルを実態に合わせて書き換えて使う。",
        "・妊活・妊娠の時期を後ろにずらすと、その後のフェーズも全部後ろにずれる。年齢と妊娠率の関係があるため、結婚式と妊活の順番は意識的に決める。",
        "・費用は目安であり、地域・式場・産院・治療内容で大きく変わる。ご祝儀・親の援助・出産育児一時金50万円・児童手当・育児休業給付金は差し引く前の金額。",
        "・制度は2026年8月時点。RSV母子免疫ワクチンの定期接種化（2026年4月〜）、こども誰でも通園制度（2026年4月〜）、出生後休業支援給付金、育児時短就業給付を反映。出産費用の保険適用化は2027〜2028年度の見込みのため、出産育児一時金50万円を前提にしている。",
        "・助成の金額・回数・申請時期は自治体差が非常に大きい。必ず住んでいる市区町村の公式情報で確認すること。",
        "・本表は計画テンプレートであり、医療上の助言ではない。健康・治療・接種の判断は主治医に、制度の適用可否は自治体・健保・勤務先・ハローワークに確認すること。",
        "・商品・サービス名は2026年8月時点で一般に流通している選択肢の例であり、特定の推奨ではない。価格と仕様は必ず最新の情報を確認すること。",
        "・③リスク管理表の確率・影響の初期値は一般的な想定値。2人の実感に合わせて必ず書き換えること。",
    ]
    for n in notes:
        c = ws.cell(row=r, column=1, value=n)
        c.font = body_font(size=9, color="4B564E")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 30
        r += 1

    set_widths(ws, {"A": 30, "B": 12, "C": 14, "D": 14, "E": 12, "F": 12, "G": 12})
    ws.sheet_view.showGridLines = False


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_roadmap(wb)
    wbs_last = sheet_wbs(wb)
    risk_last = sheet_risk(wb)
    sheet_intro(wb, wbs_last, risk_last)
    wb.active = 0
    wb.save(OUT)
    lv2 = sum(1 for x in TASKS if x[2] == 2)
    print("wrote %s" % OUT)
    print("  phases: %d / tasks: %d (lv2=%d, lv3=%d) / risks: %d"
          % (len(PHASES), len(TASKS), lv2, len(TASKS) - lv2, len(RISKS)))


if __name__ == "__main__":
    main()
